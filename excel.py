import os
import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel(datasets_data, output_path: str) -> str:
    """
    Generate a multi-sheet Excel file from dataset results.
    
    Args:
        datasets_data: Dictionary mapping dataset_id to {'columns': [...], 'rows': [...]}
        output_path: Path where the Excel file will be saved
        
    Returns:
        Path to the generated Excel file
    """
    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    if wb.active:
        wb.remove(wb.active)
    
    # Create a sheet for each dataset
    for dataset_id, data in datasets_data.items():
        columns = data.get('columns', [])
        rows = data.get('rows', [])
        
        # Create sheet name (max 31 chars for Excel limit)
        sheet_name = dataset_id[:31]
        
        # Create sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers with styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
            # Set column width based on header length
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = len(column_name) + 2
        
        # Add data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                # Auto-adjust column width for data
                column_letter = get_column_letter(col_idx)
                if cell_value:
                    cell_length = len(str(cell_value))
                    current_width = ws.column_dimensions[column_letter].width or 8
                    if cell_length + 2 > current_width:
                        ws.column_dimensions[column_letter].width = cell_length + 2
        
        # Add borders to all cells with data
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply border to header row
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.border = border
        
        # Apply border to data rows
        for row_idx in range(2, len(rows) + 2):
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
    
    # Save the workbook
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    
    return str(output_path)
